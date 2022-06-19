import numpy as np
import matplotlib.pyplot as plt
import yaml
import japanize_matplotlib

from openpyxl import Workbook
import openpyxl
import pandas
import string

import os.path
from os import path

import datetime

from tqdm import tqdm

# from keras.models import Sequential
# from keras.layers.core import Dense, Activation, Dropout

out_dir = '../output/'
filename = f'sim-log-{datetime.date.today()}'


with open('simulation/config.yaml') as file:
    params = yaml.load(file, Loader=yaml.FullLoader)

    dt =    params['dt']
    l =     params['l']
    lf =    params['lf']   
    lr = l - lf

    Cf =    params['Cf']
    Cr =    params['Cr']
    Iz =    params['Iz']
    m =     params['m']

    tp =    params['tp']

    v =     params['v']/3.6

    A = (-m*(lf*Cf - lr*Cr))/(2*l**2*Cr*Cf)

    radius = params['radius']
    straight_len = params['straight_len']

    x = params['x_init']
    y = params['y_init']

    yaw = params['yaw']

class Course():
    def __init__(self):
        theta = np.linspace(0, 2*np.pi, 40000)
        r = np.sqrt(radius**2)
        x1 = r*np.cos(theta) + straight_len
        x2 = r*np.sin(theta) + (radius+y)
        result = np.hstack((x1.reshape(-1,1),x2.reshape(-1,1)))

        xs = np.linspace(0,straight_len,1000)
        ys = np.array([10]*1000)
        lower = np.hstack((xs.reshape(-1,1),ys.reshape(-1,1)))
        final = np.concatenate((lower,result))

        self.track_all = final

class Vehicle():
    def __init__(self, x, y, v, yaw, yaw_rate, beta):
        course = Course()
        self.full_track = course.track_all

        self.count = 0 
        self.x = x
        self.y = y

        self.v = v

        self.yaw = yaw
        self.yaw_rate = yaw_rate
        self.beta = beta 

    def predict(self):
        self.gamma = (self.delta*self.v)/((1+A*self.v**2)*l)

        if self.delta == 0:
            self.sin_pre = 1
            self.cos_pre = 0
        else:
            self.sin_pre = np.sin(self.gamma*tp)/self.gamma
            self.cos_pre = (1 - np.cos(self.gamma*tp))/self.gamma
        
        self.xx_dot = self.v*self.sin_pre
        self.yy_dot = self.v*self.cos_pre

        self.x_pred = self.x + dt*(np.cos(self.yaw)*self.xx_dot - np.sin(self.yaw)*self.yy_dot)
        self.y_pred = self.y + dt*(np.sin(self.yaw)*self.xx_dot + np.cos(self.yaw)*self.yy_dot)

        self.curve = (self.beta_dot+self.yaw_rate)/self.v

        def rotate(P, theta):
            c, s = np.cos(theta), np.sin(theta)
            R = np.array(([c, s], [-s, c]), dtype=np.float)
            return P.dot(R)

        track_global = self.full_track
        track_local = rotate(track_global - (self.x, self.y), -self.yaw)
        distances = np.linalg.norm(track_local - (0,0), axis=1)
        min_distance = np.min(distances)
        min_index = np.argmin(distances)

        check = (0 - track_local[min_index-1][0])*(track_local[min_index][1]-track_local[min_index-1][1]) \
            - (0 -track_local[min_index-1][1])*(track_local[min_index][0] - track_local[min_index-1][0])

        if check < 0:
            sign = -1
        elif check > 0:
            sign = 1
        else:
            sign = 0

        feedback = sign*min_distance
        self.cte = min_distance
        return feedback

    def update(self, delta):

        def normalize_angle(angle):
            if angle > np.pi:
                angle = angle - 2 * np.pi
            if angle < -np.pi:
                angle = angle + 2 * np.pi
            return angle
        
        self.yaw = normalize_angle(self.yaw)
        self.beta = normalize_angle(self.beta)
        
        self.count += 1
        self.delta = np.clip(delta, -np.pi/4, np.pi/4)
        self.gamma = (self.delta*self.v)/((1+A*self.v**2)*l)

        if self.delta == 0:
            self.sin1 = 1
            self.cos1 = 0
        else:
            self.sin1 = np.sin(self.gamma)/self.gamma 
            self.cos1 = (1-np.cos(self.gamma))/self.gamma

        self.x_dot = (self.v)*(self.sin1)
        self.y_dot = (self.v)*(self.cos1)

        self.x += dt*(np.cos(self.yaw)*self.x_dot - np.sin(self.yaw)*self.y_dot)
        self.y += dt*(np.sin(self.yaw)*self.x_dot + np.cos(self.yaw)*self.y_dot )

        self.beta_dot = (-(Cr+Cf)/(m*self.v))*self.beta + (((Cr*lr-Cf*lf)/(m*self.v**2))-1)*self.yaw_rate + Cf*self.delta/(m*self.v)
        self.yaw_rate_dot = (Cr*lr - Cf*lf)*self.beta/Iz - (Cr*lr**2 + Cf*lf**2)*self.yaw_rate/(Iz*self.v) + Cf*lf*self.delta/Iz

        self.beta = self.beta + self.beta_dot*dt
        self.yaw_rate = self.yaw_rate + self.yaw_rate_dot*dt

        self.yaw =  self.yaw + self.yaw_rate*dt

class PID():
    def __init__(self, P=0.2, I=0.0, D=0.0, deltatime=0.01, windup_guard=20.0):
        self.Kp=P
        self.Ki=I
        self.Kd=D
        self.deltatime=deltatime
        self.windup_guard=windup_guard
        self.reset()

    def reset(self):
        self.PTerm = 0.0
        self.ITerm = 0.0
        self.DTerm = 0.0
        self.last_error = 0.0

    def update(self, feedback_value, target_value):
        error = target_value - feedback_value
        delta_error = error - self.last_error
        self.PTerm = self.Kp * error
        self.ITerm += error * self.deltatime
        self.ITerm = np.clip(self.ITerm, -self.windup_guard, self.windup_guard)

        self.DTerm = delta_error / self.deltatime
        self.last_error = error

        output = self.PTerm + (self.Ki * self.ITerm) + (self.Kd * self.DTerm)
        return output

def plot_trajectory():
    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.suptitle('走行軌跡')
    ax.set_xlabel('x 座標 [m]')
    ax.set_ylabel('y 座標 [m]')
    ax.plot(car.full_track[:,0], car.full_track[:,1], 'o', c='r',ms=0.5, alpha=0.4, label='course')
    ax.plot(X, Y, ls='dashed',c='k',lw=1.5, ms=0.7, label='path')       
    # for k,v in [(0,0),(499,5),(999,10),(1499,15),(1999,20)]:
        # ax.plot(X[k], Y[k], '^', label='$t={}s$'.format(v))
    ax.legend(fontsize=8)
    ax.grid()
    fig.tight_layout()

def plot_data():
    data =[X, Y, yaw, beta, Curvature, CTE]
    title = ['x 座標','y 座標','ヨ―角','横すべり角', '曲率','CTE']
    units = ['[m]','[m]','[rad]','[rad]','[1/m]', '[m]']
    fig, axs = plt.subplots(len(data),1,figsize=(14,8),sharex=True)
    axs = axs.flat
    axs[len(data)-1].set_xlabel('走行時間 [s]')
    for u in range(0,len(data)):
        axs[u].grid()
        axs[u].plot(t,data[u],'-', lw=0.8, c=f'C{u}')
        axs[u].set_title(title[u], fontsize=10)
        axs[u].set_ylabel(units[u])
    fig.tight_layout()

def plot_PID():
    plt.figure(figsize=(13,5))
    plt.title('PID control')
    plt.xlabel('走行時間 (s)')
    plt.ylabel('PID (PV)')
    plt.plot(t, target, label='target', lw=0.6, c='r')
    plt.plot(t, feedbacks, label='feedback', lw=0.6, c='k', ls='dashed')
    plt.plot(t, NN, label='NN')

    plt.legend(loc='lower right')
    plt.grid()

def write_csv(total_steps):
    wb = Workbook()
    ws = wb.active
    ws.title = 'log_data'
    alphabet = list(string.ascii_uppercase)
    columns_label = [
        ('time','[s]'),
        ('x','[m]'),
        ('y','[m]'),
        ('v_x','[km/h]'),
        ('v_y','[km/h]'),
        ('yaw','[rad]'),
        ('beta','[rad]'),
        ('delta','[rad]'),
        ('cte','[m]'),
        ('Curvature','[1/m]')
    ]

    for idx, (k,value) in enumerate(columns_label):
        ws[f'{alphabet[idx]}1'] = k
        ws[f'{alphabet[idx]}2'] = value

    for step in range(total_steps):
        ws[f'A{step+3}'] = f'{t[step]+0.01:.2f}'
        ws[f'B{step+3}'] = f'{X[step]:.2f}'
        ws[f'C{step+3}'] = f'{Y[step]:.2f}'
        ws[f'D{step+3}'] = f'{v*3.6:.2f}'
        ws[f'E{step+3}'] = f'{v_y[step]*3.6:.2f}'        
        ws[f'F{step+3}'] = f'{yaw[step]:.2f}'
        ws[f'G{step+3}'] = f'{beta[step]:.2f}'
        ws[f'H{step+3}'] = f'{delta[step]:.2f}'
        ws[f'I{step+3}'] = f'{CTE[step]:.2f}'
        ws[f'J{step+3}'] = f'{Curvature[step]:.2f}'
        
        


if __name__=='__main__':

    car = Vehicle(
        x = x,
        y = y,
        v = v,
        yaw = yaw,
        yaw_rate = 0,
        beta = 0.0
    )

    X, Y, yaw, beta, CTE, delta, v_y, t, Curvature,NN = [], [], [], [],[], [], [] ,[], [], []

    feedbacks, target = [], []
    feedback_value = 0.0
    pid = PID(P=1)
    
    total_steps = 5000

    from keras.models import load_model

    model = load_model('../saved_model/dbm.h5')
    nn_next = 0
    
    for i in tqdm(range(total_steps)):
        t.append(i/100)

        car.update(delta = (nn_next))
        target_value = car.predict()
        feedback_value += pid.update(feedback_value, target_value)

        test = np.zeros((1,3))
        test[:,0] = car.v*3.6
        test[:,1] = car.v*(car.beta+car.yaw)*3.6
        test[:,2] = car.curve
        
        nn_next = float(model.predict(test))

        NN.append(float(nn_next))
        feedbacks.append(feedback_value)
        target.append(target_value)    
        X.append(car.x)
        Y.append(car.y)
        v_y.append(car.v*(car.beta+car.yaw))
        yaw.append(car.yaw)
        beta.append(car.beta)
        CTE.append(car.cte)
        delta.append(car.delta)
        Curvature.append(car.curve)

    plot_PID()
    plot_data()
    plot_trajectory()

    write_csv(total_steps)


    plt.show()